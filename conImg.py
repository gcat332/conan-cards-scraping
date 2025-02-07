import requests
from PIL import Image
from io import BytesIO
import openpyxl
from openpyxl.drawing.image import Image as XLImage

# เปิดไฟล์ Excel ที่มี URL อยู่ในคอลัมน์ B
file_path = "conan_cards.xlsx"  # ระบุชื่อไฟล์ Excel
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# วนลูปอ่าน URL จากคอลัมน์ B และใส่รูปลงในคอลัมน์ A
for row in range(2, ws.max_row + 1):  # เริ่มจากแถวที่ 2 (ข้าม header)
    img_url = ws[f"B{row}"].value  # อ่าน URL จากคอลัมน์ B
    
    if img_url:  # ตรวจสอบว่ามี URL หรือไม่
        try:
            # ดาวน์โหลดรูปภาพ
            response = requests.get(img_url)
            response.raise_for_status()  # ตรวจสอบว่าดาวน์โหลดสำเร็จ

            # แปลง WebP เป็น PNG หรือ JPG
            image = Image.open(BytesIO(response.content))
            new_width = 100  # กำหนดขนาดกว้าง (pixel)
            aspect_ratio = new_width / float(image.width)
            new_height = int(float(image.height) * aspect_ratio)
            image = image.resize((new_width, new_height), Image.ANTIALIAS)

            # บันทึกเป็นไฟล์ PNG
            image_path = f"temp_image_{row}.png"
            image.save(image_path, "PNG")

            # ใส่รูปลงในคอลัมน์ A
            img = XLImage(image_path)
            ws.add_image(img, f"A{row}")

            # ปรับขนาดเซลล์ให้พอดีกับรูป
            ws.row_dimensions[row].height = new_height * 0.75
            ws.column_dimensions["A"].width = new_width * 0.14

            print(f"✅ ใส่รูปลง A{row} สำเร็จ")

        except Exception as e:
            print(f"❌ ไม่สามารถดาวน์โหลดรูปจาก {img_url}: {e}")

# บันทึกไฟล์ Excel ใหม่
output_file = "conan_cards_wp.xlsx"
wb.save(output_file)
print(f"✅ บันทึกไฟล์ Excel ใหม่: {output_file}")